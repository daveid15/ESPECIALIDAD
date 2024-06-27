from django.shortcuts import render, redirect, get_object_or_404
from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group, User
from django.core.mail import EmailMultiAlternatives
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Avg, Count, Q, Max
from django.urls import reverse, reverse_lazy
from django.utils.decorators import method_decorator
from django.conf import settings
import json
import pandas as pd
from openpyxl import Workbook
from registration.models import Profile
from inventario.models import Product
from administrator.views import validar_string, validar_nombre
import xlwt
import traceback
import os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import re
from django.http import HttpResponse



@login_required
def inventario_main(request, page=None, search=None):
    profile = Profile.objects.get(user_id=request.user.id)
    if profile.group_id != 1:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('check_group_main')

    page = request.GET.get('page', page)
    search = request.GET.get('search', search)

    if request.method == 'POST':
        search = request.POST.get('search')
        page = None

    if search is None or search == "None":
        p_list_array = Product.objects.all().order_by('supply_name')
    else:
        p_list_array = Product.objects.filter(supply_name__icontains=search).order_by('supply_name')

    # Filtrar los productos con bajo stock
    low_stock_products = []
    for p in p_list_array:
        if p.supply_total is not None and str(p.supply_total).isdigit():
            if int(p.supply_total) < 5:  # Ajusta este valor según tu criterio para "bajo stock"
                low_stock_products.append({
                    'id': p.id,
                    'supply_name': p.supply_name,
                    'supply_code': p.supply_code,
                    'supply_unit': p.supply_unit,
                    'supply_initial_stock': p.supply_initial_stock,
                    'supply_input': p.supply_input,
                    'supply_output': p.supply_output,
                    'supply_total': p.supply_total,
                    'low_stock': True,
                })
            

    paginator = Paginator(low_stock_products, 10)  # Ajusta el número de elementos por página según sea necesario
    p_list_paginate = paginator.get_page(page)

    template_name = 'inventario/inventario_main.html'
    return render(request, template_name, {
        'template_name': template_name,
        'p_list_paginate': p_list_paginate,
        'paginator': paginator,
        'page': page
    })

@login_required
def crear_producto(request):
    profile = Profile.objects.get(user_id=request.user.id)
    if profile.group_id != 1:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('check_group_main')
    #category_list = Category.objects.all().order_by('category_name')
    template_name = 'inventario/crear_producto.html'
    return render(request,template_name,{'profile':profile})#,'category_list':category_list})
#CREAR PRODUCTO TAL VEZ

"""
    producto_save Vista para guardar un nuevo producto en el sistema.

    Requiere que el usuario esté autenticado y tenga permisos de administrador (grupo_id = 1).
    Los datos del producto se reciben a través de un formulario POST y se validan antes de ser guardados en la base de datos.

    Métodos HTTP admitidos: POST

    Parámetros esperados en el formulario POST:
    - supply_name: Nombre del producto (cadena de texto).
    - supply_unit: Unidad de medida del producto ('kg' o 'LATA (330 ml)').
    - supply_code: Código único del producto (formato SKU seguido de 4 dígitos).
    - supply_initial_stock: Stock inicial del producto (número entero).
    - supply_input: Cantidad de ingreso de stock del producto (opcional, entero).
    - supply_output: Cantidad de salida de stock del producto (opcional, entero).

    Validaciones realizadas:
    - Todos los campos deben estar completos.
    - El nombre del producto debe ser válido según la función validar_nombre.
    - La unidad de medida debe ser válida ('kg' o 'LATA (330 ml)').
    - El stock inicial debe ser un número entero.
    - El código del producto debe tener el formato SKU seguido de 4 dígitos.
    - Si no se proporcionan supply_input o supply_output, se consideran como 0.

    Retorna:
    - Redirige a 'inventario_main' si el producto se guarda correctamente.
    - Redirige a 'crear_producto' con un mensaje de error si hay problemas en el formulario o si el método de envío no es POST.
    - Redirige a 'check_group_main' si el usuario no tiene permisos de administrador (grupo_id != 1).
"""
@login_required
def producto_save(request):
    profile = Profile.objects.get(user_id=request.user.id)
    if profile.group_id != 1:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una área para la que no tiene permisos')
        return redirect('check_group_main')
    if request.method == 'POST':

        supply_name = request.POST.get('supply_name')
        supply_unit = request.POST.get('supply_unit')
        supply_code = request.POST.get('supply_code')
        supply_initial_stock = request.POST.get('supply_initial_stock')
        supply_input = request.POST.get('supply_input')
        supply_output = request.POST.get('supply_output')
        
        supply_code = f'SKU{supply_code}'
   
        if supply_name == '' or supply_unit == '' or supply_initial_stock == '' or supply_input == '' or supply_output == '':
            messages.add_message(request, messages.INFO, 'Debes ingresar toda la información')
            return redirect('crear_producto')
        
        if not validar_nombre(supply_name, request):
            messages.add_message(request, messages.INFO, 'Debes ingresar un nombre válido')
            return redirect('crear_producto')
        
        if supply_unit not in ['kg', 'LATA (330 ml)']:
            messages.add_message(request, messages.INFO, 'La unidad no es válida')
            return redirect('crear_producto')
        
        if not supply_initial_stock.isdigit():
            messages.add_message(request, messages.INFO, 'El stock inicial debe ser un número entero')
            return redirect('crear_producto')
        
        if not re.match(r'^SKU\d{4}$', supply_code):
            messages.add_message(request, messages.INFO, 'El código del producto debe tener el formato SKU seguido de 4 dígitos')
            return redirect('crear_producto')
        
        if supply_input is None:
            supply_input = 0
        if supply_output is None:
            supply_output = 0

        producto_save = Product(
            supply_name=supply_name,
            supply_code=supply_code,
            supply_unit=supply_unit,
            supply_initial_stock=supply_initial_stock,
            supply_input=int(supply_input),
            supply_output=int(supply_output),
            supply_total=int(supply_initial_stock) + int(supply_input) - int(supply_output),
)
        producto_save.save()
        messages.add_message(request, messages.INFO, 'Producto ingresado con éxito')
        return redirect('inventario_main')
    else:
        messages.add_message(request, messages.INFO, 'Error en el método de envío')
        return redirect('check_group_main')


   #VER PRODUCTO Y se usa para editar
@login_required
def producto_ver(request, product_id):
    profile = Profile.objects.get(user_id=request.user.id)
    if profile.group_id != 1:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('check_group_main')
    product_data = Product.objects.get(pk=product_id)
    supply_code = re.sub(r'\D', '', product_data.supply_code)
    #category_list = Category.objects.all()
    #selected_category = product_data.product_category
    template_name = 'inventario/producto_ver.html'
    return render(request, template_name, {'profile': profile, 'product_data': product_data, 'supply_code':supply_code}) #'category_list': category_list, 'selected_category': selected_category})



@login_required
def producto_list(request, page=None, search=None):
    profile = Profile.objects.get(user_id=request.user.id)
    if profile.group_id != 1:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('check_group_main')

    if page is None:
        page = request.GET.get('page')
    else:
        page = page

    if request.GET.get('page') is None:
        page = page
    else:
        page = request.GET.get('page')

    if search is None:
        search = request.GET.get('search')
    else:
        search = search

    if request.GET.get('search') is None:
        search = search
    else:
        search = request.GET.get('search')

    if request.method == 'POST':
        search = request.POST.get('search')
        page = None

    if search is None or search == "None":
        p_count = Product.objects.filter(supply_name='a').count()
        p_list_array = Product.objects.all().order_by('supply_name')
    else:
        p_count = Product.objects.filter(supply_name='a').filter(supply_name__icontains=search).count()
        p_list_array = Product.objects.all().filter(supply_name__icontains=search).order_by('supply_name')

    p_list = []
    for p in p_list_array:
        p_list.append({
            'id': p.id,
            'supply_name': p.supply_name,
            'supply_code': p.supply_code,
            'supply_unit': p.supply_unit,
            'supply_initial_stock': p.supply_initial_stock,
            'supply_input': p.supply_input,
            'supply_output': p.supply_output,
            'supply_total': p.supply_total,
            #'category_name': p.product_category.category_name, # aquí se obtiene el category_name asociado al producto
        })

    paginator = Paginator(p_list, 5)
    p_list_paginate = paginator.get_page(page)

    template_name = 'inventario/producto_list.html'
    return render(request, template_name, {'template_name': template_name, 'p_list_paginate': p_list_paginate, 'paginator': paginator, 'page': page})

"""
   producto_edit Vista para editar un producto existente en el sistema.

    Requiere que el usuario esté autenticado y tenga permisos de administrador (grupo_id = 1).
    Los datos del producto se recuperan a través de su ID y se actualizan con los nuevos datos proporcionados en un formulario POST.

    Métodos HTTP admitidos: GET (para mostrar el formulario de edición) y POST (para procesar la edición).

    Parámetros de URL:
    - product_id: ID único del producto que se desea editar.

    Parámetros esperados en el formulario POST:
    - supply_name: Nombre del producto (cadena de texto).
    - supply_code: Código único del producto (formato SKU seguido de 4 dígitos).
    - supply_unit: Unidad de medida del producto ('kg' o 'LATA (330 ml)').
    - supply_initial_stock: Stock inicial del producto (número entero).
    - supply_input: Cantidad de ingreso de stock del producto (opcional, entero).
    - supply_output: Cantidad de salida de stock del producto (opcional, entero).

    Validaciones realizadas:
    - Todos los campos deben estar completos.
    - El nombre del producto debe ser válido según la función validar_nombre.
    - La unidad de medida debe ser válida ('kg' o 'LATA (330 ml)').
    - El stock inicial, entrada y salida deben ser números enteros.
    - El código del producto debe tener el formato SKU seguido de 4 dígitos.
    - Si no se proporcionan supply_input o supply_output, se consideran como 0.

    Acciones:
    - Recupera el producto existente mediante su ID.
    - Actualiza los campos del producto con los nuevos datos proporcionados.
    - Calcula el stock total del producto en base a los valores de stock inicial, entrada y salida.
    - Guarda los cambios en la base de datos y redirige a la página de detalle del producto con un mensaje de éxito.

    Redirecciones:
    - Redirige a la página de detalle del producto ('producto_ver') después de guardar los cambios correctamente.
    - Redirige a la página de verificación de grupo principal ('check_group_main') si el usuario no tiene permisos de administrador (grupo_id != 1).
"""
@login_required
def producto_edit(request, product_id):
    profile = Profile.objects.get(user_id=request.user.id)
    if profile.group_id != 1:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('check_group_main')
    
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        supply_name = request.POST.get('supply_name')
        supply_code = request.POST.get('supply_code')
        supply_unit = request.POST.get('supply_unit')
        supply_initial_stock = request.POST.get('supply_initial_stock')
        supply_input = request.POST.get('supply_input')
        supply_output = request.POST.get('supply_output')
        
        supply_code = f'SKU{supply_code}'
   
        if supply_name == '' or supply_unit == '' or supply_initial_stock == '' or supply_input == '' or supply_output == '':
            messages.add_message(request, messages.INFO, 'Debes ingresar toda la información')
            return redirect('producto_ver', product_id=product.id)
        
        if not validar_nombre(supply_name, request):
            messages.add_message(request, messages.INFO, 'Debes ingresar un nombre de producto válido')
            return redirect('producto_ver', product_id=product.id)
        
        if supply_unit not in ['kg', 'LATA (330 ml)']:
            messages.add_message(request, messages.INFO, 'La unidad no es válida')
            return redirect('producto_ver', product_id=product.id)
        
        if not supply_initial_stock.isdigit() or not supply_input.isdigit() or not supply_output.isdigit():
            messages.add_message(request, messages.INFO, 'El stock inicial, entrada y salida deben ser números enteros')
            return redirect('producto_ver', product_id=product.id)
        
        if not re.match(r'^SKU\d{4}$', supply_code):
            messages.add_message(request, messages.INFO, 'El código del producto debe tener el formato SKU seguido de 4 dígitos')
            return redirect('producto_ver', product_id=product.id)
        
        if supply_input is None:
            supply_input = 0
        if supply_output is None:
            supply_output = 0
        
        product.supply_name = supply_name
        product.supply_code = supply_code
        product.supply_unit = supply_unit
        product.supply_input = supply_input
        product.supply_output = supply_output

        # Comprobar si es la primera ediciÃ³n
        if product.supply_total == product.supply_initial_stock:
            product.supply_initial_stock = supply_initial_stock
        else:
            product.supply_initial_stock = product.supply_total

        # Calcular supply_total en base a supply_initial_stock, supply_input y supply_output
        if supply_input == "0" and supply_output == "0":
            product.supply_total = product.supply_initial_stock
        else:
            if supply_input == "0":
                product.supply_total = int(product.supply_initial_stock) - int(supply_output)
            elif supply_output == "0":
                product.supply_total = int(product.supply_initial_stock) + int(supply_input)
            else:
                product.supply_total = int(product.supply_initial_stock) + int(supply_input) - int(supply_output)

        product.save()
        messages.success(request, 'Producto editado correctamente')
        
        return redirect('producto_ver', product_id=product.id)
    else:
        product_data = Product.objects.get(pk=product_id)
        template_name = 'inventario/producto_ver.html'
        
        return render(request, template_name, {'product_data': product_data})

"""
   producto_delete Vista para eliminar un producto del sistema.

    Requiere que el usuario esté autenticado y tenga permisos de administrador (grupo_id = 1).
    Recibe el ID del producto que se desea eliminar como parámetro.

    Métodos HTTP admitidos: POST.

    Parámetros de URL:
    - product_id: ID único del producto que se desea eliminar.

    Acciones:
    - Verifica que el usuario tenga permisos de administrador (grupo_id = 1).
    - Recupera el producto mediante su ID o muestra un error 404 si no existe.
    - Elimina el producto de la base de datos.
    - Muestra un mensaje de éxito indicando que el producto ha sido eliminado correctamente.
    - Redirige a la lista de productos ('producto_list') después de eliminar el producto.

    Redirecciones:
    - Redirige a la lista de productos ('producto_list') después de eliminar el producto correctamente.
    - Redirige a la página de verificación de grupo principal ('check_group_main') si el usuario no tiene permisos de administrador (grupo_id != 1).
"""
#ELIMINAR PRODUCTO
def producto_delete(request, product_id):
    profile = Profile.objects.get(user_id=request.user.id)
    if profile.group_id != 1:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
        return redirect('check_group_main')
    product = get_object_or_404(Product, id=product_id)
    product.delete()
    messages.success(request, 'Producto eliminado correctamente')
    return redirect(reverse('producto_list'))
"""
    carga_masiva_producto Vista para mostrar el formulario de carga masiva de productos.

    Requiere que el usuario esté autenticado y tenga permisos de administrador (grupo_id = 1).

    Métodos HTTP admitidos: GET.

    Acciones:
    - Verifica que el usuario tenga permisos de administrador (grupo_id = 1).
    - Muestra el formulario de carga masiva de productos ('carga_masiva_producto.html').

    Redirecciones:
    - Redirige a la página de verificación de grupo principal ('check_group_main') si el usuario no tiene permisos de administrador (grupo_id != 1).
"""

#CARGA MASIVA   
@login_required
def carga_masiva_producto(request):
    profile = Profile.objects.get(user_id=request.user.id)
    if profile.group_id != 1:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una área para la que no tiene permisos')
        return redirect('check_group_main')
    template_name = 'inventario/carga_masiva_producto.html'
    return render(request, template_name, {'profiles': profile})
"""
    import_file_producto Vista para importar un archivo Excel de productos.

    Requiere que el usuario esté autenticado y tenga permisos de administrador (grupo_id = 1).
    Descarga un archivo de ejemplo para la importación de productos.

    Métodos HTTP admitidos: GET.

    Acciones:
    - Verifica que el usuario tenga permisos de administrador (grupo_id = 1).
    - Descarga un archivo Excel de ejemplo prellenado con datos de ejemplo para la importación de productos.

    Redirecciones:
    - Redirige a la página de verificación de grupo principal ('check_group_main') si el usuario no tiene permisos de administrador (grupo_id != 1).
"""
@login_required
def import_file_producto(request):
    profiles = Profile.objects.get(user_id=request.user.id)
    if profiles.group_id != 1:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a una área para la que no tiene permisos')
        return redirect('check_group_main')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="archivo_importacion_producto.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = 'carga_masiva'

    columns = ['Producto', 'Código', 'Unidad', 'Stock inicial']
    ws.append(columns)

    example_data = ['ej: Palta', 'ej: SKU1111', 'ej: kg/LATA (330 ml)', 'ej: 10']
    ws.append(example_data)

    wb.save(response)
    return response  
"""
   carga_masiva_producto_save Vista para procesar y guardar los datos importados de productos desde un archivo Excel.

    Requiere que el usuario esté autenticado y tenga permisos de administrador (grupo_id = 1).
    Recibe un archivo Excel de productos para ser procesado y guardado en la base de datos.

    Métodos HTTP admitidos: POST.

    Parámetros de formulario:
    - myfile: Archivo Excel que contiene los datos de productos a importar.

    Acciones:
    - Verifica que el usuario tenga permisos de administrador (grupo_id = 1).
    - Lee el archivo Excel enviado por el usuario.
    - Itera sobre los datos del archivo Excel, validando y creando cada producto.
    - Muestra un mensaje indicando la cantidad de registros importados correctamente.
    - En caso de error, muestra un mensaje de error con detalles.

    Redirecciones:
    - Redirige de vuelta a la página de carga masiva de productos ('carga_masiva_producto') después de procesar los datos correctamente o en caso de error.
"""
@login_required
def carga_masiva_producto_save(request):
    profiles = Profile.objects.get(user_id=request.user.id)
    if profiles.group_id != 1:
        messages.add_message(request, messages.INFO, 'Intenta ingresar a un área para la que no tiene permisos')
        return redirect('check_group_main')

    if request.method == 'POST':
        try:
            data = pd.read_excel(request.FILES['myfile'], engine='openpyxl', skiprows=1)
            df = pd.DataFrame(data)
            acc = 0
            for item in df.itertuples():
                supply_name = str(item[1])
                supply_code = str(item[2])
                supply_unit = str(item[3])
                supply_initial_stock = int(item[4])
                supply_input=0
                supply_output=0

                 #Valida que supply_name solo sean letras
                if not supply_name.replace('', '').isalpha():
                    raise ValueError('El nombre del insumo solo puede contener letras')

                # Valida que supply_unit solo sean las opciones kg o LATA (330 ml)
                if supply_unit not in ['kg', 'LATA (330 ml)']:
                    raise ValueError('La unidad del suministro solo puede ser "kg" o "LATA (330 ml)"')

                # Valida que supply_initial_stock solo sean números
                if not isinstance(supply_initial_stock, int):
                    raise ValueError('El stock inicial debe ser un valor numérico')
            
                producto_save = Product(
                    supply_name=supply_name,
                    supply_code=supply_code,
                    supply_unit=supply_unit,
                    supply_initial_stock=supply_initial_stock,
                    supply_output=supply_output,
                    supply_input=supply_input,
                    supply_total=supply_initial_stock,
                    
                )
                producto_save.save()
                acc += 1

            messages.add_message(request, messages.INFO, f'Carga masiva finalizada, se importaron {acc} registros')
            return redirect('carga_masiva_producto')
        except Exception as e:
            messages.add_message(request, messages.ERROR, f'Error al procesar el archivo: {str(e)}')
            return redirect('carga_masiva_producto')

    # Si la solicitud no es POST, redirige al usuario a alguna otra vista
    return redirect('carga_masiva_producto')  # Esto puede ser cambiado dependiendo de tus necesidades
    
@login_required
def descarga_reporte_producto(request):
    try:    
        profiles = Profile.objects.get(user_id = request.user.id)
        if profiles.group_id != 1:
         messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
         return redirect('check_group_main')
    
        style_2 = xlwt.easyxf('font: name Time New Roman, color-index black; font: bold on')

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        response=HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ListaProductos.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Productos')

        row_num = 0
        columns = ['Producto', 'Código', 'Unidad']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num] ,font_style)

        productos = Product.objects.all().order_by('supply_name')

        for row in productos:
            row_num += 1
            for col_num in range(4):
                if col_num == 0:
                    ws.write(row_num, col_num, row.supply_name, style_2)
                if col_num == 1:
                    ws.write(row_num, col_num, row.supply_code, style_2)   
                if col_num == 2:
                    ws.write(row_num, col_num, row.supply_unit, style_2)
             #  if col_num == 3:
            #       ws.write(row_num, col_num, str(row.product_category), style_2)

        
        wb.save(response)
        return response
    except Exception:
            traceback.print_exc()
            messages.add_message(request, messages.ERROR, 'Se produjo un error al generar el archivo Excel. Por favor, inténtelo de nuevo más tarde.')
            return redirect('producto_list')

@login_required
def reportes_main_productos(request):
    try:
        profiles = Profile.objects.get(user_id = request.user.id)
        if profiles.group_id != 1:
            messages.add_message(request, messages.INFO, 'Intenta ingresar a una area para la que no tiene permisos')
            return redirect('check_group_main')
        
        template_name = 'inventario/reportes_main_productos.html'
        return render(request,template_name)
    except:
        messages.add_message(request, messages.INFO, 'Error al acceder al pagina de reportes')
        return redirect('producto_list')


@login_required
def reporte_producto_filtro(request):
    try:
        profiles = Profile.objects.get(user_id=request.user.id)
        if profiles.group_id != 1:
            messages.add_message(request, messages.INFO, 'Intenta ingresar a un área para la que no tiene permisos')
            return redirect('check_group_main')

        if request.method == 'POST':
            producto = request.POST.get('supply_name')
            if len(producto) == 0:
                messages.add_message(request, messages.INFO, 'El producto no puede estar vacío')
                return redirect('reportes_main_productos')
            producto_count = Product.objects.filter(estado='Activo').filter(supply_name__icontains=producto).count()
            if producto_count < 1:
                messages.add_message(request, messages.INFO, 'No existe producto con la cadena buscada')
                return redirect('reportes_main')

            style_1 = xlwt.easyxf('font: name Times New Roman, color-index black, bold on', num_format_str='#,##0.00')
            style_2 = xlwt.easyxf('font: name Times New Roman, color-index black, bold on')
            

            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="ReporteProductos.xls"'
            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet('Productos')

            row_num = 0
            columns = ['Nombre', 'Precio', 'Estado'] #le borre categoria
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], style_2)

            productos_array = Product.objects.filter(estado='Activo').filter(supply_name__icontains=producto)
            for row in productos_array:
                row_num += 1
                for col_num in range(4):
                    if col_num == 0:
                        ws.write(row_num, col_num, row.supply_name, style_2)
                    if col_num == 1:
                        ws.write(row_num, col_num, row.supply_code, style_2)
                    if col_num == 2:
                        ws.write(row_num, col_num, row.supply_unit, style_2)
                    #if col_num == 3:
                    #   ws.write(row_num, col_num, str(row.product_category), style_2)

            wb.save(response)
            return response
        else:
            messages.add_message(request, messages.INFO, 'Error')
            return redirect('check_group_main')
    except:
        messages.add_message(request, messages.INFO, 'Error al generar el reporte')
        return redirect('reportes_main_productos')